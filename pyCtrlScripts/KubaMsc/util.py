# MUST CHANGE FILE NAME
import numpy as np
import re
import os
from OF_caseClass import OpenFOAMCase
import matplotlib.pyplot as plt
from pathlib import Path
import openpyxl as pxl
import math
import itertools
import logging

logging.basicConfig(level=logging.INFO, format="%(levelname)10s | %(message)s")
WHITE = "\033[0m"
BOLD = "\033[1m"
CYAN = "\033[96m"
PINK = "\033[95m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
RED = "\033[91m"


def log(func, col, msg):
    func(col + msg + WHITE)

logging.basicConfig(level=logging.INFO)
def loadDataOLD(filePath: str, tempInC: bool = False) -> np.ndarray:
    """Reads data from filePath and returns it as a numpy array. The file is assumed to have one float per line."""
    a = np.loadtxt(filePath).transpose()
    if tempInC:
        for i in range(len(a)):
            a[i] += 273.15
    return a


def loadExperimentsOLD() -> np.ndarray:
    """Loads all experimental data from the ../expData directory and returns a 2D numpy array. Assumes that each file contains one column of data. Sort the files in the directory by name, so that they are in the same order as the corresponding simulation probe locations."""
    expData = np.array(
        [
            loadDataOLD("../expData/" + series, True)
            for series in sorted(os.listdir("../expData"))
        ]
    )
    return expData


class OpcuaData:
    def __init__(self):
        self.times: list = []
        self.temps: dict = {}

    def unpack(self) -> list:
        return [self.times, self.temps]


def loadOpcuaLog(filePath: str, tempInC: bool = False) -> OpcuaData:
    """Reads data from an excel file with columns labeled 'Timestamp' for times, and 'TempX', where 'X' is a number for temperatures"""
    log(logging.info, BOLD, f"reading opcua log at {filePath}...")
    try:
        sheet = pxl.open(filePath)["Sheet1"]
    except FileNotFoundError:
        log(
            logging.critical,
            RED,
            f"Could not find specified experimental data log, please check the adress.\nCurrent adress is {filePath}",
        )
    col: int = 1
    data = OpcuaData()

    while sheet.cell(1, col)._value != None:
        row: int = 2
        header = sheet.cell(1, col)._value
        log(logging.info, CYAN, f"reading column {header}")
        if header == "Timestamp":
            startTime = sheet.cell(2, col)._value
            while sheet.cell(row, col)._value != None:
                data.times.append((float(sheet.cell(row, col)._value) - startTime) * 60)
                row += 1
            if len(data.times) > 0:
                log(logging.info, GREEN, "successfully read time column")
            else:
                log(logging.critical, RED, "could not read time column!")
            logging.debug(str(data.times))
        elif "Temp" in header:
            temps = []
            while sheet.cell(row, col)._value != None:
                temps.append(float(sheet.cell(row, col)._value) + 273.15)
                row += 1
            num = str(re.fullmatch(r"Temp(\d+)", header).group(1))
            data.temps[num] = temps
            if len(data.temps[num]) > 0:
                log(logging.info, GREEN, f"successfully read temp column number {num}")
                logging.debug(str(data.temps[num]))
            else:
                log(
                    logging.critical, RED, f"temp column number {num} could not be read"
                )
        else:
            log(logging.warning, YELLOW, f"undefined column id '{header}', skipping")
        col += 1
    log(logging.info, BOLD, f"finished reading opcua log at {filePath}")
    return data


def probeSim(field: str, locs: list[tuple], case: OpenFOAMCase) -> None:
    """Probes the simulation at the specified locations with OpenFOAM probe utility. The locations should be in the format "(x y z)"."""
    log(logging.info, BOLD, "Starting probe process...")
    locsString = "("
    for i in range(len(locs)):
        locsString += locs[i][1]
    locsString += ")"
    log(logging.debug, WHITE, "Generated locsString")
    logging.debug(locsString)
    print(PINK)
    case.replace(
        [
            [
                "system/probeField",
                ["locationPlaceholder", "fieldPlaceholder"],
                [locsString, field],
            ]
        ]
    )
    case.runCommands(["postProcess -func probeField -dict system/probeField>log.probe"])
    case.replace(
        [
            [
                "system/probeField",
                [locsString, field],
                ["locationPlaceholder", "fieldPlaceholder"],
            ]
        ]
    )
    print(WHITE)
    try:
        data = np.loadtxt(
            case.dir + "/postProcessing/probeField/0/" + field,
            skiprows=7,  # IMPORTANT!!! this should be probe count +1, temporarily set directly
        ).transpose()
    except FileNotFoundError:
        log(
            logging.critical,
            RED,
            f"The probe failed to generate the appropriate file, searched in {case.dir}/postProcessing/probeField/0/",
        )

    log(logging.info, GREEN, "probe successfully generated data")
    log(logging.info, CYAN, "compiling probe data")
    dataOut = OpcuaData()
    time, data = data[0], data[1:]
    for i in range(len(data)):
        series = locs[i][0]
        dataOut.temps[f"{series}"] = data[i]
    dataOut.times = time
    log(logging.info, GREEN, "data successfully compiled")
    logging.debug(dataOut)
    log(logging.info, BOLD, "ending probe")
    return dataOut


def linInterpol(a: float, b: float, ratio: float) -> float:
    """Linear interpolation between a and b with ratio. ratio should be between 0 and 1."""
    return a - (a - b) * ratio


def trimDataOLD(expData: np.ndarray, simData: np.ndarray) -> np.ndarray:
    """Trims the longer of expData and simData to match the length of the shorter one using linear interpolation. Assumes start and end times to be the same for both."""
    if len(simData[0] != len(expData[0])):
        expLong: bool = len(simData[0]) < len(expData[0])
        if expLong:
            data = [simData, expData]
        else:
            data = [expData, simData]
        dataShortened = np.zeros([len(simData), min(len(simData[0]), len(expData[0]))])

        for probe in range(len(data[1])):
            nLong = len(data[1][probe])
            nShort = len(data[0][probe])
            dataTemp = np.zeros([len(data[0][probe])])
            dataTemp[0] = data[1][probe][0]
            for datI in range(1, len(dataTemp) - 1):
                dataTemp[datI] = linInterpol(
                    data[1][probe][int(np.floor(datI * (nLong / nShort)))],
                    data[1][probe][int(np.floor(datI * (nLong / nShort)) + 1)],
                    (datI * (nLong / nShort)) - np.floor(datI * (nLong / nShort)),
                )
            dataTemp[-1] = data[1][probe][-1]
            dataShortened[probe] = dataTemp
        data[1] = dataShortened
        if expLong:
            return [data[1], data[0]]
        else:
            return data
    else:
        return [expData, simData]


def trimDataNEW(expData: dict, simData: dict):
    log(logging.info, BOLD, "trimming data to matching length")
    key1 = list(expData.keys())[0]
    if len(simData[key1]) != len(expData.get(key1)):
        expLong: bool = len(simData.get(key1)) < len(expData.get(key1))
        if expLong:
            log(logging.debug, WHITE, "expData is longer than simData")
            data = [simData, expData]
        else:
            log(logging.debug, WHITE, "simData is longer than expData")
            data = [expData, simData]
        nShort = len(data[0].get(key1))
        nLong = len(data[1].get(key1))
        log(logging.debug, WHITE, f"{nLong} VS {nShort} entries")

        for seriesN in data[0].keys():
            log(logging.info, CYAN, f"trimming data series {seriesN}")
            tempData = np.zeros([len(data[0][str(seriesN)])])
            longSeries = data[1][str(seriesN)]
            tempData[0] = longSeries[0]
            ratio = (nLong - 1) / nShort

            for datI in range(1, len(tempData) - 1):
                tempData[datI] = linInterpol(
                    longSeries[int(np.floor(datI * ratio))],
                    longSeries[int(np.floor(datI * ratio) + 1)],
                    (datI * ratio) - np.floor(datI * ratio),
                )
            tempData[-1] = longSeries[-1]
            log(logging.info, GREEN, f"series {seriesN} succesfully trimmed")
            logging.debug(f"result: {tempData}")
            data[1][str(seriesN)] = tempData
        log(logging.info, BOLD, "finished trimming data")
        if expLong:
            return [data[1], data[0]]
        else:
            return data
    else:
        return [expData, simData]


def plotResults(expData: np.ndarray, simData: np.ndarray):
    xs = range(len(expData[0]))
    for i in enumerate(simData):
        plt.plot(xs, expData[i[0]], xs, i[1])
        plt.show()


def plotDicts(
    time, simDict, expDict, label1="Simulation", label2="Experiment", figsize_per_plot=5
):
    log(logging.info, BOLD, "plotting data")
    if simDict.keys() != expDict.keys():
        raise ValueError("dict1 and dict2 must have identical keys.")

    keys = list(simDict.keys())
    n = len(keys)

    ncols = math.ceil(math.sqrt(n))
    nrows = math.ceil(n / ncols)

    logging.debug(f"with {ncols} columns and {nrows} rows")

    fig, axes = plt.subplots(
        nrows, ncols, figsize=(figsize_per_plot * ncols, 4 * nrows)
    )

    axes = axes.flatten()

    for ax, key in zip(axes, keys):
        ax.plot(time, simDict[key], label=label1)
        ax.plot(time, expDict[key], label=label2)
        ax.set_title(str(key))
        ax.set_xlabel("Time")
        ax.set_ylabel("Value")
        ax.grid(True)
        ax.legend()
    for ax in axes[n:]:
        ax.set_visible(False)

    plt.tight_layout()
    log(logging.info, GREEN, f"showing {ncols*nrows} plots")
    plt.show()
    log(logging.info, GREEN, f"saving {ncols*nrows} plots")
    plt.savefig("plots/test.png")
    log(logging.info, BOLD, "plotting finished")


def transplantTime(case: OpenFOAMCase, realEndTime: float) -> None:
    case.replace(
        [["system/controlDict", ["ENDTIMEPLACEHOLDER"], [str(math.floor(realEndTime))]]]
    )


def runSingleSim(baseCaseDir: str, targetDir: str, runSim: bool):
    print(PINK)
    case = OpenFOAMCase()
    case.loadOFCaseFromBaseCase("../../tutorials/" + baseCaseDir)
    case.changeOFCaseDir("../../ZZ_cases" + targetDir)
    print(WHITE)
    expData: OpcuaData = loadOpcuaLog("../temp/opcua_log_20260211_150638.xlsx")
    dur = expData.times[-1] - expData.times[0]
    if runSim:
        print(PINK)
        case.copyBaseCase()
        transplantTime(case, dur)
        case.runCommands(["./Allclean", "sbatch --wait ./AllrunSlurm"])
    simData: OpcuaData = probeSim(
        "T",
        [
            (5, "(0.576 0 0.320)"),
            (6, "(0.057 0.071 0.595)"),
            (7, "(0.536 0.097 0.580)"),
            (8, "(0.056 0.074 0.058)"),
            (3, "(0.544 0.082 0.062)"),
            (4, "(0.024 0.150 0.572)"),
        ],
        case,
    )
    times, simData = simData.unpack()
    expData = expData.unpack()[1]
    expData, simData = trimDataNEW(expData, simData)
    plotDicts(times, simData, expData)


def getSetsFromDict(dictIn: dict) -> tuple[tuple]:
    len1 = len(dictIn.get(list(dictIn.keys())[0]))
    vals = []
    for values in dictIn.values():
        vals.append(values)
    sets = itertools.product(*vals)
    return tuple(sets)


def runMultiSim(baseCaseDir: str, seriesName: str, parDict: dict):
    parSets: tuple[tuple] = getSetsFromDict(parDict)
    for simN, pars in enumerate(parSets):
        parString = ""
        for par in pars:
            parString += str(par) + "_"
        parString += "out"
        target: str = "/00_breads/" + seriesName + "/" + parString
        # change pars
        # runSingleSim


runSingleSim("ovenTest", "/00_breads/lolPar", True)
# loadOpcuaLog("../expData/opcua_log_20260211_115313.xlsx")

# exp = {"1": [1, 2, 3, 4, 5], "2": [10, 20, 30, 40, 50]}
# sim = {
#     "1": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
#     "2": [10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
# }
# data = trimDataNEW(exp, sim)
# print(data[0], "\n/////////////////////////////////////////\n", data[1])

# pars = {"L": [1, 2, 3], "T": [10, 20, 30]}
# runMultiSim("a", "hhh", pars)
